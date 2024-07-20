#!/usr/bin/env python
# coding: utf-8

# In[1]:


# Import packages
import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.expected_conditions import presence_of_element_located

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# In[2]:


# Load workbook data
wb = load_workbook(filename='input.xlsx')
sheet = wb.active

# In[3]:


# Create array of destinations extracted from the A2-A15 in Excel
destinations = []
for cell in list(range(2, 16)):
    destinations.append(sheet[f"A{cell}"].value)


# In[4]:


# Helper functions

def click_train_btn(driver: webdriver.Chrome, wait: WebDriverWait):
    ## Find train btn
    wait.until(
        presence_of_element_located(
            (By.XPATH, '/html/body/div[2]/div[3]/div[8]/div[3]/div[1]/div[2]/div/div[2]/div/div/div/div[3]/button')))
    btn = driver.find_elements_by_xpath(
        '/html/body/div[2]/div[3]/div[8]/div[3]/div[1]/div[2]/div/div[2]/div/div/div/div[3]/button')[0]
    ## Click train btn
    btn.click()


def scrape_duration(wait: WebDriverWait):
    ## Get time duration in 1st suggested trip, by using the XPATH of the 1st time element
    d1 = wait.until(
        presence_of_element_located((By.XPATH, '//*[@id="section-directions-trip-0"]/div[1]/div/div[1]/div')))

    return d1.text


# In[5]:


# Start web scraping
print("Starting Chrome....")

with webdriver.Chrome() as driver:
    startTime = time.time()
    wait = WebDriverWait(driver, 10)

    col_num = 2  # starting from column B in excel

    while True:
        col = get_column_letter(col_num)
        origin = sheet[f"{col}1"].value  # Get origin
        if origin != None:
            for i in range(len(destinations)):
                directions_url = f"https://www.google.com/maps/dir/{origin}/{destinations[i]}"
                driver.get(directions_url)  # Open directions url

                click_train_btn(driver, wait)
                time.sleep(0.1)  # wait for new trip results after clicking train btn

                duration = scrape_duration(wait)
                print(f"{directions_url}: {duration}")

                # Write duration into excel
                sheet[col + str(i + 2)] = duration
        else:
            break
        col_num += 1
    print("Total time spent scraping is ", time.time() - startTime, "secs")
    wb.save("./output.xlsx")

